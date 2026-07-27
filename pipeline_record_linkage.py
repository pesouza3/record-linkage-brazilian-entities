import pandas as pd
import numpy as np
import os
import re
import time
import gc  # Coletor de lixo para limpar a memória RAM
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from unidecode import unidecode

# Bibliotecas de Machine Learning e Record Linkage
import recordlinkage
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from xgboost import XGBClassifier
from sklearn.metrics import f1_score, precision_score, recall_score, confusion_matrix, precision_recall_curve

# Bibliotecas para formatação executiva do Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Início do monitoramento temporal do pipeline
tempo_inicio_geral = time.time()

# ======================================================================
# CONFIGURAÇÃO DE DIRETÓRIOS E ISOLAMENTO TEMPORAL DE PASTAS
# ======================================================================
DIRETORIO_PAI = '/Users/pedrosouza/Desktop/MBA/TCC/'
ARQUIVO_CSV_LOCAL = os.path.join(DIRETORIO_PAI, 'empresas.csv')

# Identificador exclusivo de timestamp para a rodada
carimbo_tempo = datetime.now().strftime("execucao_%Y_%m_%d_%H%M%S")
DIRETORIO_SESSAO = os.path.join(DIRETORIO_PAI, carimbo_tempo)

if not os.path.exists(DIRETORIO_SESSAO):
    os.makedirs(DIRETORIO_SESSAO)

# Definição dos caminhos de destino dentro da pasta da sessão
ARQUIVO_BASE_PARQUET = os.path.join(DIRETORIO_SESSAO, 'base_100k_bruta.parquet')
ARQUIVO_FEATURES_TMP = os.path.join(DIRETORIO_SESSAO, 'features_temporarias.parquet')
ARQUIVO_EXCEL_AMOSTRAS = os.path.join(DIRETORIO_SESSAO, 'amostras_record_linkage_tcc.xlsx')
CAMINHO_FIGURA_2 = os.path.join(DIRETORIO_SESSAO, 'figura2_curva_pr.png')
CAMINHO_FIGURA_3 = os.path.join(DIRETORIO_SESSAO, 'figura3_matriz_confusao.png')
CAMINHO_FIGURA_4 = os.path.join(DIRETORIO_SESSAO, 'figura4_importancia_features.png')

print("\n" + "="*70)
print(f"  DIRETÓRIO DA RODADA LOCAL PROTEGIDA: {carimbo_tempo}")
print("="*70)

# ======================================================================
# PASSO 1: CARREGAMENTO DE ALTA VELOCIDADE VIA ARQUIVO CSV LOCAL
# ======================================================================
print("\n" + "="*70)
print("  PASSO 1: CARREGAMENTO DA AMOSTRA COMPLETA (100.000 REGISTROS)")
print("="*70)
t_etapa = time.time()

# Leitura direta do arquivo local
df_bruto = pd.read_csv(ARQUIVO_CSV_LOCAL, nrows=100000, usecols=['cnpj', 'razao_social', 'uf'])
df_bruto = df_bruto.dropna(subset=['cnpj', 'razao_social', 'uf']).reset_index(drop=True)

# Geração das estatísticas descritivas para a TABELA 1 do TCC
n_empresas = len(df_bruto)
media_tokens = df_bruto['razao_social'].apply(lambda x: len(str(x).split())).mean()
print(f"[TABELA 1 - DATA] Número total de empresas extraídas localmente: {n_empresas:,}")
print(f"[TABELA 1 - DATA] Média de tokens por razão social: {media_tokens:.2f}")

# Salvamento de auditoria em parquet
df_bruto.to_parquet(ARQUIVO_BASE_PARQUET, compression='snappy')

# Padronização primitiva de tipos
df_bruto['cnpj'] = df_bruto['cnpj'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(14)
df_bruto['razao_social'] = df_bruto['razao_social'].astype(str)
df_bruto['uf'] = df_bruto['uf'].astype(str).str.upper()
df_bruto['id'] = df_bruto.index + 1
print(f"[TEMPO] Carregamento local concluído em {time.time() - t_etapa:.2f} segundos.")

# ======================================================================
# PASSO 2 E 3: PRÉ-PROCESSAMENTO E INJEÇÃO DE RUÍDOS CONTROLADOS
# ======================================================================
print("\n" + "="*70)
print("  PASSO 2 e 3: PRÉ-PROCESSAMENTO E SIMULAÇÃO DE RUÍDO cadastral")
print("="*70)

def limpar_texto_nativo(texto):
    texto = unidecode(str(texto)).lower()
    sufixos_regex = r'\b(ltda|s\.a\.|sa|s/a|me|eireli|epp|meireli)\b'
    texto = re.sub(sufixos_regex, '', texto)
    texto = re.sub(r'[^a-z0-9\s]', '', texto)
    return " ".join(texto.split())

df_bruto['nome_limpo'] = df_bruto['razao_social'].apply(limpar_texto_nativo)
df_bruto = df_bruto[df_bruto['nome_limpo'] != ""].reset_index(drop=True)
df_bruto['id'] = df_bruto.index + 1

# Mantido o Blocking por tetragrama [:4] + UF para blindar o produto cartesiano
df_bruto['bloco_memoria'] = df_bruto['nome_limpo'].str[:4].str.strip() + "_" + df_bruto['uf']
df_base = df_bruto[['id', 'cnpj', 'razao_social', 'nome_limpo', 'bloco_memoria']].copy()

df_comprada = df_base.copy()
contagem_ruidos = {'intacto': 0, 'typo': 0, 'omissao': 0, 'inversao': 0}

def injetar_ruido_estruturado(texto):
    if len(texto) < 7: 
        contagem_ruidos['intacto'] += 1
        return texto
    escolha = np.random.choice(['intacto', 'typo', 'omissao', 'inversao'], p=[0.70, 0.10, 0.10, 0.10])
    contagem_ruidos[escolha] += 1
    palavras = texto.split()
    if escolha == 'typo':
        return texto[:-2] + 'x' + texto[-1:]
    elif escolha == 'omissao':
        return texto[:-4]
    elif escolha == 'inversao' and len(palavras) >= 2:
        return " ".join([palavras[-1]] + palavras[:-1])
    return texto

df_comprada['razao_social_ruidosa'] = df_comprada['razao_social'].apply(injetar_ruido_estruturado)
df_comprada['nome_limpo'] = df_comprada['razao_social_ruidosa'].apply(limpar_texto_nativo)
df_comprada['bloco_memoria'] = df_comprada['nome_limpo'].str[:4].str.strip() + "_" + df_comprada['bloco_memoria'].str.split('_').str[-1]

print(f"[TABELA 2 - DATA] Distribuição volumétrica dos ruídos injetados:")
print(contagem_ruidos)

# ======================================================================
# PASSO 4: INDEXAÇÃO DE ALTA PERFORMANCE (BLOCKING OTIMIZADO)
# ======================================================================
print("\n" + "="*70)
print("  PASSO 4: INDEXAÇÃO DE ALTA PERFORMANCE (BLOCKING OTIMIZADO)")
print("="*70)
t_etapa = time.time()

df_base = df_base.set_index('id')
df_comprada = df_comprada.set_index('id')

indexer = recordlinkage.Index()
indexer.block('bloco_memoria')
pares_candidatos = indexer.index(df_base, df_comprada)
total_pares = len(pares_candidatos)
print(f"[TABELA 4 - DATA] Mapeamento reduzido com segurança para {total_pares:,} pares candidatos.")
print(f"[TEMPO] Blocagem executada com sucesso em {time.time() - t_etapa:.2f} segundos.")

# ======================================================================
# PASSO 5: ENGENHARIA DE ATRIBUTOS COM SALVAMENTO EM DISCO EM STRINGS
# ======================================================================
print("\n" + "="*70)
print("  PASSO 5: CÁLCULO DE SIMILARIDADE COM STREAMING ANTI-TRAVAMENTO (SSD)")
print("="*70)
t_etapa = time.time()

comparador = recordlinkage.Compare()
comparador.string('nome_limpo', 'nome_limpo', method='jarowinkler', label='sim_nome_jw')
comparador.string('nome_limpo', 'nome_limpo', method='levenshtein', label='sim_nome_lv')

# Processamento fatiado em blocos estáveis de 500 mil pares por vez
tamanho_lote = 500000
primeiro_lote = True

for i in range(0, total_pares, tamanho_lote):
    sub_par = pares_candidatos[i:i+tamanho_lote]
    sub_feature = comparador.compute(sub_par, df_base, df_comprada)
    
    # Aplicação do Ground Truth (CNPJ) no pedaço da RAM
    c_base = df_base.loc[sub_feature.index.get_level_values(0), 'cnpj'].values
    c_comp = df_comprada.loc[sub_feature.index.get_level_values(1), 'cnpj'].values
    sub_feature['is_match'] = np.where(c_base == c_comp, 1, 0)
    
    # Dump direto para disco limpando a RAM de forma contínua
    if primeiro_lote:
        sub_feature.to_parquet(ARQUIVO_FEATURES_TMP, engine='pyarrow', index=True)
        primeiro_lote = False
    else:
        df_existente = pd.read_parquet(ARQUIVO_FEATURES_TMP)
        df_consolidado = pd.concat([df_existente, sub_feature])
        df_consolidado.to_parquet(ARQUIVO_FEATURES_TMP, engine='pyarrow', index=True)
        del df_existente, df_consolidado
        
    del sub_feature
    gc.collect() # Liberação forçada da memória cache
    
    progresso = min(100, (i + tamanho_lote) * 100 // total_pares)
    print(f"[PROGRESSO DISCO] Calculando e descarregando no SSD: {progresso}% concluído...")

print("[STATUS] Resgatando matriz unificada estável do SSD para alimentar o ML...")
features = pd.read_parquet(ARQUIVO_FEATURES_TMP)

if os.path.exists(ARQUIVO_FEATURES_TMP):
    os.remove(ARQUIVO_FEATURES_TMP)
print(f"[TEMPO] Engenharia de atributos concluída em {time.time() - t_etapa:.2f} segundos.")

# ======================================================================
# PASSO 6 E 7: ENGENHARIA DE MACHINE LEARNING & PREDIÇÃO DE SCORES
# ======================================================================
print("\n" + "="*70)
print("  PASSO 6 e 7: TREINAMENTO MULTI-ALGORITMOS E MATRIZES GRÁFICAS")
print("="*70)

X = features[['sim_nome_jw', 'sim_nome_lv']]
y = features['is_match']

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42, stratify=y)
proporcao_desbalanceamento = (len(y_train) - sum(y_train)) / sum(y_train)

modelos = {
    'Regressão Logística': LogisticRegression(),
    'Random Forest': RandomForestClassifier(n_estimators=50, random_state=42, n_jobs=-1),
    'XGBoost': XGBClassifier(n_estimators=50, max_depth=4, scale_pos_weight=proporcao_desbalanceamento, random_state=42, eval_metric='logloss')
}

resultados_tabela3 = []
y_pred_baseline = np.where(X_test['sim_nome_jw'] > 0.85, 1, 0)
resultados_tabela3.append({
    'Modelo': 'Baseline (Regra Jaro-Winkler > 0.85)',
    'Precisão': precision_score(y_test, y_pred_baseline, zero_division=0),
    'Revocação': recall_score(y_test, y_pred_baseline, zero_division=0),
    'F1-Score': f1_score(y_test, y_pred_baseline, zero_division=0)
})

plt.figure(figsize=(8, 6))
for nome, model in modelos.items():
    print(f"[ML] Ajustando classificador: {nome}...")
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    y_proba = model.predict_proba(X_test)[:, 1]
    
    resultados_tabela3.append({
        'Modelo': nome,
        'Precisão': precision_score(y_test, y_pred, zero_division=0),
        'Revocação': recall_score(y_test, y_pred, zero_division=0),
        'F1-Score': f1_score(y_test, y_pred, zero_division=0)
    })
    p_curve, r_curve, _ = precision_recall_curve(y_test, y_proba)
    plt.plot(r_curve, p_curve, label=nome)

plt.xlabel('Revocação (Recall)')
plt.ylabel('Precisão (Precision)')
plt.title('Figura 2: Curva Precisão-Revocação dos Modelos Comparados')
plt.legend()
plt.grid(True)
plt.savefig(CAMINHO_FIGURA_2, dpi=300)
plt.close()

df_tabela3 = pd.DataFrame(resultados_tabela3)
print("\n--- [TABELA 3] COMPARATIVO DE DESEMPENHO DOS MODELOS ---")
print(df_tabela3.to_string(index=False))

melhor_modelo = modelos['Random Forest']
y_pred_rf = melhor_modelo.predict(X_test)
cm = confusion_matrix(y_test, y_pred_rf)

plt.figure(figsize=(6, 5))
sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', cbar=False, xticklabels=['Não Match', 'Match'], yticklabels=['Não Match', 'Match'])
plt.ylabel('Realidade Analítica (Filtro CNPJ)')
plt.xlabel('Predição Operational do Modelo (ML)')
plt.title('Figura 3: Matriz de Confusão do Classificador Random Forest')
plt.tight_layout()
plt.savefig(CAMINHO_FIGURA_3, dpi=300)
plt.close()

importancias = melhor_modelo.feature_importances_
features_nomes = X.columns
plt.figure(figsize=(6, 4))
sns.barplot(x=importancias, y=features_nomes, hue=features_nomes, palette='viridis', legend=False)
plt.title('Figura 4: Grau de Importance Relativa dos Atributos')
plt.xlabel('Ganhos de Informação (Feature Importance)')
plt.tight_layout()
plt.savefig(CAMINHO_FIGURA_4, dpi=300)
plt.close()

# ======================================================================
# PASSO 8: GARIMPO E SALVAMENTO DA PLANILHA EXCEL DE AMOSTRAS REAIS
# ======================================================================
print("\n" + "="*70)
print("  PASSO 8: EXTRAÇÃO DE AMOSTRAS E SALVAMENTO PLANILHA EXCEL")
print("="*70)

features['predicao_rf'] = melhor_modelo.predict(X)
ex_matches = features[features['is_match'] == 1].head(8)
ex_nao_matches = features[(features['is_match'] == 0) & (features['sim_nome_jw'] > 0.75)].head(4)
amostra_indices = pd.concat([ex_matches, ex_nao_matches])

registros_excel = []
for idx, row in amostra_indices.iterrows():
    id_a, id_b = idx
    registros_excel.append({
        "Razão Social (Base A)": df_base.loc[id_a, 'razao_social'],
        "Razão Social Ruidosa (Base B)": df_comprada.loc[id_b, 'razao_social_ruidosa'],
        "Jaro-Winkler": row['sim_nome_jw'],
        "Levenshtein": row['sim_nome_lv'],
        "Predição RF": "Match" if row['predicao_rf'] == 1 else "Não Match",
        "Validação CNPJ": "Match" if row['is_match'] == 1 else "Não Match"
    })

df_excel_final = pd.DataFrame(registros_excel)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Amostras TCC"
ws.views.sheetView[0].showGridLines = True

header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid")
title_font = Font(name="Arial", size=13, bold=True, color="1F385C")
data_font = Font(name="Arial", size=10)

match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") 
no_match_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

ws.merge_cells("A1:F1")
ws["A1"] = "USP Esalq - MBA em Data Science & Analytics | Carga Local Otimizada de Alta Velocidade"
ws["A1"].font = Font(name="Arial", size=9, italic=True, color="595959")
ws.merge_cells("A2:F2")
ws["A2"] = f"TCC: Casos Práticos de Pareamento | {carimbo_tempo}"
ws["A2"].font = title_font

ws.append([]) 
headers_col = list(df_excel_final.columns)
ws.append(headers_col)

for col_num, h_text in enumerate(headers_col, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left", vertical="center")
    cell.border = thin_border

for row_idx, r_data in enumerate(df_excel_final.values, start=5):
    ws.append(list(r_data))
    is_match_rf = r_data[4] == "Match"
    for col_idx in range(1, len(headers_col) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        if col_idx in [3, 4]: 
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx in [5, 6]: 
            cell.fill = match_fill if is_match_rf else no_match_fill
            cell.font = Font(name="Arial", size=10, bold=True, color="375623" if is_match_rf else "C65911")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            if row_idx % 2 == 0: cell.fill = zebra_fill

for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 3)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(ARQUIVO_EXCEL_AMOSTRAS)
print(f"[SUCESSO] Planilha salva em: '{ARQUIVO_EXCEL_AMOSTRAS}'")
print(f"[FIM] Experimento finalizado com sucesso e segurança em {time.time() - tempo_inicio_geral:.2f} segundos.\n")
